"""Read Demography_Refrences.xlsx using openpyxl (no pandas needed)."""
import openpyxl

REF_PATH = r"c:\Users\vaibh\Desktop\CKA Projects\Chronical\Demography\Original Data\Demography_Refrences.xlsx"

wb = openpyxl.load_workbook(REF_PATH)
print("Sheets:", wb.sheetnames)

for shname in wb.sheetnames:
    ws = wb[shname]
    print(f"\n\n=== Sheet: '{shname}' ===")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # skip fully empty rows
        if all(c is None for c in row):
            continue
        print(f"  Row {i+1}: {[str(c)[:80] if c is not None else '' for c in row]}")
