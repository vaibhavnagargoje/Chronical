"""
One-shot script to extract each data sheet from MH_health.xlsx into a separate CSV file.
Sheets that are metadata-only (Heading_Shortenings_Index, Notes_Index) are skipped.
Sheet names with spaces/parentheses are normalised to underscores for filenames,
but the (2) suffix is preserved as _2 so the reference sheet mappings stay intact.
"""
import os
import csv
import openpyxl

XLSX_PATH = os.path.join(os.path.dirname(__file__), 'MH_health.xlsx')
OUTPUT_DIR = os.path.dirname(__file__)

# These are metadata/index sheets, not data sheets
SKIP_SHEETS = {'Heading_Shortenings_Index', 'Notes_Index'}


def sheet_to_csv_filename(sheet_name: str) -> str:
    """
    Convert a sheet name like 'HMIS_IV (2)' → 'hmis_iv_2.csv'
    and 'DSA_FamilyWelfarePrograms' → 'dsa_familywelfareprograms.csv'
    """
    name = sheet_name.strip()
    name = name.replace(' (2)', '_2')   # preserve the _2 suffix
    name = name.replace(' ', '_')
    name = name.lower()
    return f"{name}.csv"


def extract_all():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            print(f"  [SKIP] {sheet_name} (metadata)")
            continue

        ws = wb[sheet_name]
        csv_name = sheet_to_csv_filename(sheet_name)
        csv_path = os.path.join(OUTPUT_DIR, csv_name)

        # Read the header row — filter out None columns (unused cols in Excel)
        raw_header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        # Find the last non-None column index
        last_col = 0
        for i, h in enumerate(raw_header):
            if h is not None:
                last_col = i
        header = raw_header[: last_col + 1]
        num_cols = len(header)

        rows_written = 0
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(header)

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                row_data = list(row[:num_cols])
                # Skip completely empty rows
                if all(v is None for v in row_data):
                    continue
                writer.writerow(row_data)
                rows_written += 1

        print(f"  [OK] {csv_name} — {rows_written} rows, {num_cols} columns")

    print("\nDone!")


if __name__ == '__main__':
    extract_all()
